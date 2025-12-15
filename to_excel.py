#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel出力モジュール: 構造化データをExcelファイルに変換
"""

import openpyxl
from openpyxl.styles import Font, Alignment
import logging

# ロギング設定
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Excelのカラム順序（仕様に基づく15個のカラム）
EXCEL_COLUMNS = [
    '班',
    '氏名',
    '住所',
    'TEL',
    'メールアドレス',
    '事務局',
    '会計',
    '書記',
    '名簿',
    '防犯防災',
    '回覧広報',
    '地域コミュ',
    '環境美化',
    '厚生福祉',
    '補足事項'
]


def create_excel(data_list, output_path):
    """
    データリストからExcelファイルを作成

    Args:
        data_list (list): 班長データの辞書のリスト
        output_path (str): 出力Excelファイルのパス

    Returns:
        str: 作成されたExcelファイルのパス
    """
    try:
        # 新しいワークブックを作成
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "班長役員希望内訳"

        # ヘッダー行を作成
        logger.info("Excelヘッダーを作成中...")
        for col_idx, col_name in enumerate(EXCEL_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            # ヘッダーを太字にする
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # データ行を追加
        logger.info(f"{len(data_list)} 件のデータを書き込み中...")
        for row_idx, data in enumerate(data_list, 2):  # 2行目から開始（1行目はヘッダー）
            for col_idx, col_name in enumerate(EXCEL_COLUMNS, 1):
                value = data.get(col_name, 'null')
                # 'null'文字列はそのまま保持（仕様に基づく）
                ws.cell(row=row_idx, column=col_idx, value=value)

        # 列幅を自動調整
        logger.info("列幅を調整中...")
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # 最大50文字まで
            ws.column_dimensions[column].width = adjusted_width

        # Excelファイルを保存
        wb.save(output_path)
        logger.info(f"Excelファイルを保存しました: {output_path}")

        return output_path

    except Exception as e:
        logger.error(f"Excel作成中にエラーが発生しました: {str(e)}")
        raise


def create_csv(data_list, output_path):
    """
    データリストからCSVファイルを作成（UTF-8エンコーディング）

    Args:
        data_list (list): 班長データの辞書のリスト
        output_path (str): 出力CSVファイルのパス

    Returns:
        str: 作成されたCSVファイルのパス
    """
    try:
        import csv

        logger.info(f"CSVファイルを作成中: {output_path}")

        # UTF-8エンコーディングでCSVファイルを作成
        with open(output_path, 'w', encoding='utf-8', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=EXCEL_COLUMNS)

            # ヘッダーを書き込み
            writer.writeheader()

            # データを書き込み
            for data in data_list:
                # EXCEL_COLUMNSの順序でデータを書き込み
                row_data = {col: data.get(col, 'null') for col in EXCEL_COLUMNS}
                writer.writerow(row_data)

        logger.info(f"CSVファイルを保存しました: {output_path}")
        return output_path

    except Exception as e:
        logger.error(f"CSV作成中にエラーが発生しました: {str(e)}")
        raise


if __name__ == "__main__":
    # テスト用コード
    import sys
    import json

    if len(sys.argv) < 3:
        print("使用方法: python to_excel.py <入力JSONファイル> <出力Excelファイル>")
        print("例: python to_excel.py data.json output.xlsx")
        sys.exit(1)

    json_file = sys.argv[1]
    output_file = sys.argv[2]

    try:
        # JSONファイルからデータを読み込み
        with open(json_file, 'r', encoding='utf-8') as f:
            data = json.load(f)

        # Excelファイルを作成
        if output_file.endswith('.xlsx'):
            create_excel(data, output_file)
        elif output_file.endswith('.csv'):
            create_csv(data, output_file)
        else:
            print("エラー: 出力ファイルは.xlsxまたは.csv形式である必要があります")
            sys.exit(1)

        print(f"ファイルを作成しました: {output_file}")

    except Exception as e:
        print(f"エラー: {str(e)}")
        sys.exit(1)
